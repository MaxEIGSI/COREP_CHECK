from __future__ import annotations

from dataclasses import dataclass
from itertools import product
from pathlib import Path
import ast
import math
import re
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import openpyxl
import pandas as pd

try:
    from v2.get_RC_value import (
        ALL_SENTINEL,
        DEFAULT_BASED_TEMPLATE_PATH,
        DEFAULT_BASED_TEMPLATE_SHEET,
        DEFAULT_COREP_DIR,
        DEFAULT_MAPPING_TABLE_PATH,
        build_column_code_map,
        build_row_code_map,
        contains_template_hint,
        load_table_sheet_mapping,
        letter_to_sequence,
        normalize_axis_code,
        parse_selector,
        sequence_to_marker,
        worksheet_to_dataframe,
    )
except ModuleNotFoundError:
    from v2.get_RC_value import (
        ALL_SENTINEL,
        DEFAULT_BASED_TEMPLATE_PATH,
        DEFAULT_BASED_TEMPLATE_SHEET,
        DEFAULT_COREP_DIR,
        DEFAULT_MAPPING_TABLE_PATH,
        build_column_code_map,
        build_row_code_map,
        contains_template_hint,
        load_table_sheet_mapping,
        letter_to_sequence,
        normalize_axis_code,
        parse_selector,
        sequence_to_marker,
        worksheet_to_dataframe,
    )

INTERVAL_TOLERANCE = 1e-6
AGG_FUNCS = {"sum", "count", "max", "min"}


@dataclass(frozen=True)
class Coordinate:
    template: str
    table: str
    row: Optional[str]
    column: Optional[str]
    sheet: str


@dataclass
class RuleDetail:
    coordinates: Tuple[str, str, Optional[str], Optional[str], str]
    expected: Any
    actual: Any
    passed: bool
    message: str = ""
    formula_values: Optional[Dict[str, Any]] = None
    precondition_values: Optional[Dict[str, Any]] = None


@dataclass
class RuleResult:
    rule_id: str
    status: str
    details: List[RuleDetail]
    reason: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "rule_id": self.rule_id,
            "status": self.status,
            "reason": self.reason,
            "details": [
                {
                    "coordinates": detail.coordinates,
                    "expected": detail.expected,
                    "actual": detail.actual,
                    "passed": detail.passed,
                    "message": detail.message,
                    "formula_values": detail.formula_values or {},
                    "precondition_values": detail.precondition_values or {},
                }
                for detail in self.details
            ],
        }


@dataclass(frozen=True)
class RefSpec:
    template: Optional[str]
    table: Optional[str]
    row: Optional[str]
    row_alias: Optional[str]
    column: Optional[str]
    column_alias: Optional[str]
    sheet: Optional[str]
    include_sheet_pattern: Optional[str]
    exclude_sheets: Tuple[str, ...]
    wildcard: bool
    label: str


@dataclass
class ParsedExpression:
    text: str
    refs: Dict[str, RefSpec]
    ast_root: ast.AST


@dataclass
class SheetContext:
    dataframe: pd.DataFrame
    row_map: Dict[str, int]
    col_map: Dict[str, Any]


class RuleEngineError(Exception):
    pass


class NonDslRuleError(RuleEngineError):
    pass


def normalize_template_id(template_code: str) -> str:
    """
    Accept standard codes (C07.00, F18.00, J01.00, P02.04)
    and extended variants (F01.01_dp, F04.03.1, C17.01_v2, …).
    Only the base two-part prefix (Letter + nn.nn) is required.
    """
    raw = str(template_code).strip().replace(" ", "")
    upper = raw.upper()
    if not re.match(r"[A-Z]\d{2}\.\d{2}", upper):
        raise RuleEngineError(f"Invalid template code format: {template_code}")
    return upper


def template_to_file_name(template_code: str) -> str:
    norm = normalize_template_id(template_code)
    # Remove dots to build compact code, lowercase any _XX suffixes (e.g. _DP → _dp)
    compact = norm.replace(".", "")
    compact = re.sub(r"_([A-Z]+)$", lambda m: "_" + m.group(1).lower(), compact)
    return f"G_EU_C_{compact}.xlsx"


def template_to_file_path(template_code: str, corep_dir: Path) -> Path:
    path = corep_dir / template_to_file_name(template_code)
    if not path.exists():
        raise RuleEngineError(f"COREP file not found for template {template_code}: {path}")
    return path


def split_table_identifier(table_name: str) -> Tuple[str, Optional[str]]:
    cleaned = str(table_name).strip().upper().replace(" ", "")
    parts = cleaned.split(".")
    if len(parts) >= 3 and parts[-1].isalpha():
        return ".".join(parts[:-1]), parts[-1].lower()
    return cleaned, None


def resolve_sheet_for_table_generic(
    wb: openpyxl.Workbook,
    template_code: str,
    table_name: str,
) -> str:
    table_template, suffix_letter = split_table_identifier(table_name)
    norm_template = normalize_template_id(template_code)

    if normalize_template_id(table_template) != norm_template:
        raise RuleEngineError(f"Table {table_name} does not match template {template_code}")
    if suffix_letter is None:
        raise RuleEngineError(f"Table {table_name} has no suffix letter")

    marker = sequence_to_marker(letter_to_sequence(suffix_letter))
    marker_pattern = re.compile(rf"-\s*{marker}\b")

    for ws in wb.worksheets:
        has_marker = False
        has_template = False
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                if marker_pattern.search(text):
                    has_marker = True
                if contains_template_hint(text, norm_template):
                    has_template = True
                if has_marker and has_template:
                    return ws.title

    raise RuleEngineError(f"No worksheet found for table {table_name} (marker {marker})")


def _normalize_sheet_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value).upper())


def _normalize_table_key(value: Any) -> str:
    return str(value).strip().upper().replace(" ", "")


def _table_template_key(table_key: str) -> Optional[str]:
    match = re.match(r"([A-Z]\d{2}\.\d{2})", table_key)
    return match.group(1) if match else None


def _find_mapped_sheet_value(mapping: Dict[str, str], table_key: str) -> Optional[str]:
    direct = mapping.get(table_key)
    if direct is not None:
        return direct

    template_key = _table_template_key(table_key)
    if template_key is None:
        return None

    candidate_outputs = {
        output
        for key, output in mapping.items()
        if _table_template_key(key) == template_key
    }
    if len(candidate_outputs) == 1:
        return next(iter(candidate_outputs))

    return mapping.get(template_key)


def _match_sheet_name(wb: openpyxl.Workbook, mapped_name: str) -> Optional[str]:
    if mapped_name in wb.sheetnames:
        return mapped_name

    target = _normalize_sheet_key(mapped_name)
    if not target:
        return None

    for sheet in wb.sheetnames:
        if _normalize_sheet_key(sheet) == target:
            return sheet

    for sheet in wb.sheetnames:
        candidate = _normalize_sheet_key(sheet)
        if candidate.startswith(target) or target in candidate:
            return sheet

    return None


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def to_number(value: Any) -> Optional[float]:
    if is_empty(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip()
    try:
        return float(text)
    except ValueError:
        return None


def flatten(values: Iterable[Any]) -> List[Any]:
    out: List[Any] = []
    for value in values:
        if isinstance(value, list):
            out.extend(flatten(value))
        elif isinstance(value, tuple):
            out.extend(flatten(list(value)))
        else:
            out.append(value)
    return out


def values_sum(values: Iterable[Any]) -> Optional[float]:
    nums = [to_number(v) for v in flatten(values)]
    nums = [n for n in nums if n is not None]
    if not nums:
        return None
    return float(sum(nums))


def values_largest_sum(count_value: Any, values: Iterable[Any]) -> Optional[float]:
    count_num = to_number(count_value)
    if count_num is None:
        return None

    top_n = max(0, int(count_num))
    nums = [to_number(v) for v in flatten(values)]
    nums = [n for n in nums if n is not None]
    if not nums:
        return None
    if top_n == 0:
        return 0.0

    largest = sorted(nums, reverse=True)[:top_n]
    return float(sum(largest))


def values_count(values: Iterable[Any]) -> int:
    return sum(1 for value in flatten(values) if not is_empty(value))


def values_max(values: Iterable[Any]) -> Optional[float]:
    nums = [to_number(v) for v in flatten(values)]
    nums = [n for n in nums if n is not None]
    if not nums:
        return None
    return max(nums)


def values_min(values: Iterable[Any]) -> Optional[float]:
    nums = [to_number(v) for v in flatten(values)]
    nums = [n for n in nums if n is not None]
    if not nums:
        return None
    return min(nums)


def values_length(value: Any) -> int:
    if is_empty(value):
        return 0
    return len(str(value).strip())


def split_selector_groups(value: Any) -> Optional[List[str]]:
    if value is None:
        return None

    if isinstance(value, float) and pd.isna(value):
        return None

    if not isinstance(value, str):
        return parse_selector(value)

    text = value.strip()
    if not text:
        return None
    if text.lower() == "all":
        return [ALL_SENTINEL]

    parts: List[str] = []
    current: List[str] = []
    brace_depth = 0

    for char in text:
        if char == "{":
            brace_depth += 1
        elif char == "}" and brace_depth > 0:
            brace_depth -= 1

        if char == ";" and brace_depth == 0:
            chunk = "".join(current).strip()
            if chunk:
                parts.append(chunk)
            current = []
            continue

        current.append(char)

    tail = "".join(current).strip()
    if tail:
        parts.append(tail)

    return parts or None


def parse_axis_assignment_groups(value: Any, axis_prefix: str) -> Optional[List[Dict[str, str]]]:
    tokens = split_selector_groups(value)
    if not tokens or ALL_SENTINEL in tokens:
        return None

    groups: List[Dict[str, str]] = []
    normalized_prefix = axis_prefix.lower()

    for token in tokens:
        text = str(token).strip()
        if not text:
            continue

        inner = text[1:-1].strip() if text.startswith("{") and text.endswith("}") else text
        if "=" not in inner:
            continue

        group: Dict[str, str] = {}
        for part in inner.split(";"):
            item = part.strip()
            if not item or "=" not in item:
                continue

            alias_name, raw_value = [chunk.strip() for chunk in item.split("=", 1)]
            alias_key = alias_name.lower()
            if not alias_key.startswith(normalized_prefix):
                continue

            axis_value = normalize_axis_code(raw_value)
            if axis_value is None:
                continue
            group[alias_key] = axis_value.zfill(4)

        if group:
            groups.append(group)

    return groups or None


def combine_alias_groups(
    row_groups: Optional[List[Dict[str, str]]],
    column_groups: Optional[List[Dict[str, str]]],
) -> List[Dict[str, str]]:
    if row_groups and column_groups:
        return [{**row_map, **col_map} for row_map, col_map in product(row_groups, column_groups)]
    if row_groups:
        return row_groups
    if column_groups:
        return column_groups
    return [{}]


class CorepDataRepository:
    def __init__(
        self,
        corep_dir: str | Path = DEFAULT_COREP_DIR,
        mapping_table_path: str | Path = DEFAULT_MAPPING_TABLE_PATH,
    ):
        self.corep_dir = Path(corep_dir)
        self.mapping_table_path = Path(mapping_table_path)
        self._workbooks: Dict[str, openpyxl.Workbook] = {}
        self._sheet_context: Dict[Tuple[str, str], SheetContext] = {}
        self._table_sheet_mapping = load_table_sheet_mapping(self.mapping_table_path)

    def workbook_for_template(self, template: str) -> openpyxl.Workbook:
        norm = normalize_template_id(template)
        if norm not in self._workbooks:
            self._workbooks[norm] = openpyxl.load_workbook(
                template_to_file_path(norm, self.corep_dir), data_only=True
            )
        return self._workbooks[norm]

    def get_table_sheet(self, template: str, table: str) -> str:
        mapped_sheet = _find_mapped_sheet_value(self._table_sheet_mapping, _normalize_table_key(table))
        if mapped_sheet is not None:
            wb = self.workbook_for_template(template)
            matched = _match_sheet_name(wb, mapped_sheet)
            if matched is not None:
                return matched

        return resolve_sheet_for_table_generic(self.workbook_for_template(template), template, table)

    def all_sheets(self, template: str) -> List[str]:
        return [ws.title for ws in self.workbook_for_template(template).worksheets]

    def context(self, template: str, sheet: str) -> SheetContext:
        key = (normalize_template_id(template), sheet)
        if key not in self._sheet_context:
            ws = self.workbook_for_template(template)[sheet]
            df = worksheet_to_dataframe(ws)
            row_map = {k.zfill(4): v for k, v in build_row_code_map(df).items()}
            col_map = {k.zfill(4): v for k, v in build_column_code_map(df).items()}
            self._sheet_context[key] = SheetContext(df, row_map, col_map)
        return self._sheet_context[key]


class InMemoryCorepDataRepository:
    def __init__(self, data_mapping: Mapping[str, Any]):
        self._sheet_context: Dict[Tuple[str, str], SheetContext] = {}
        self._template_sheets: Dict[str, List[str]] = {}
        self._table_sheets: Dict[Tuple[str, str], str] = {}

        raw_table_map = data_mapping.get("table_to_sheet", {}) if isinstance(data_mapping, Mapping) else {}
        for key, value in raw_table_map.items():
            if isinstance(key, tuple) and len(key) == 2:
                template, table = key
            else:
                parts = str(key).split("|")
                if len(parts) != 2:
                    continue
                template, table = parts[0], parts[1]
            self._table_sheets[(normalize_template_id(str(template)), str(table).upper())] = str(value)

        nested = data_mapping.get("frames") if isinstance(data_mapping, Mapping) and "frames" in data_mapping else data_mapping

        if not isinstance(nested, Mapping):
            raise RuleEngineError("data_mapping must be a mapping")

        for template_key, by_sheet in nested.items():
            template = normalize_template_id(str(template_key))
            if not isinstance(by_sheet, Mapping):
                continue
            self._template_sheets.setdefault(template, [])

            for sheet_name, frame in by_sheet.items():
                if not isinstance(frame, pd.DataFrame):
                    continue
                sheet = str(sheet_name)
                self._template_sheets[template].append(sheet)
                row_map = {k.zfill(4): v for k, v in build_row_code_map(frame).items()}
                col_map = {k.zfill(4): v for k, v in build_column_code_map(frame).items()}
                self._sheet_context[(template, sheet)] = SheetContext(frame, row_map, col_map)

            self._template_sheets[template] = sorted(set(self._template_sheets[template]))

    def get_table_sheet(self, template: str, table: str) -> str:
        key = (normalize_template_id(template), str(table).upper())
        if key in self._table_sheets:
            return self._table_sheets[key]

        available = self.all_sheets(template)
        if len(available) == 1:
            return available[0]

        raise RuleEngineError(
            f"Cannot resolve table->sheet for {table} in template {template}. "
            "Provide data_mapping['table_to_sheet']."
        )

    def all_sheets(self, template: str) -> List[str]:
        norm = normalize_template_id(template)
        sheets = self._template_sheets.get(norm, [])
        if not sheets:
            raise RuleEngineError(f"No sheets found for template {template}")
        return sheets

    def context(self, template: str, sheet: str) -> SheetContext:
        key = (normalize_template_id(template), str(sheet))
        if key not in self._sheet_context:
            raise RuleEngineError(f"No in-memory dataframe for {key[0]} / {key[1]}")
        return self._sheet_context[key]


class DimensionResolver:
    def __init__(self, repository: Any):
        self.repository = repository

    def _parse_templates(self, value: Any) -> List[str]:
        tokens = parse_selector(value) or []
        return [normalize_template_id(token) for token in tokens if token != ALL_SENTINEL]

    def _parse_tables(self, value: Any, template: str) -> List[str]:
        tokens = parse_selector(value)
        if not tokens:
            return [template]

        out: List[str] = []
        for token in tokens:
            normalized = str(token).strip().upper()
            if normalized == ALL_SENTINEL:
                continue
            if re.fullmatch(r"[A-Z]\d{2}\.\d{2}\.[A-Z]+", normalized) and normalized.startswith(template):
                out.append(normalized)

        return out or [template]

    def _parse_axis(self, value: Any) -> Optional[List[str]]:
        tokens = parse_selector(value)
        if not tokens:
            return None
        if ALL_SENTINEL in tokens:
            return None

        out: List[str] = []
        for token in tokens:
            token_text = str(token).strip()
            if "=" in token_text:
                continue
            norm = normalize_axis_code(token)
            if norm is not None:
                out.append(norm.zfill(4))

        return out or None

    def _parse_sheets(self, value: Any) -> Optional[List[str]]:
        tokens = parse_selector(value)
        if not tokens:
            return None
        if ALL_SENTINEL in tokens:
            return None
        return [str(token).strip() for token in tokens if str(token).strip()]

    def resolve_scope(self, rule_row: pd.Series) -> Dict[str, Any]:
        templates = self._parse_templates(rule_row.get("Templates used"))
        if not templates:
            raise RuleEngineError("No valid template in rule")

        rows = self._parse_axis(rule_row.get("Rows"))
        row_alias_groups = parse_axis_assignment_groups(rule_row.get("Rows"), "r")
        columns = self._parse_axis(rule_row.get("Columns"))
        column_alias_groups = parse_axis_assignment_groups(rule_row.get("Columns"), "c")
        sheets_override = self._parse_sheets(rule_row.get("Sheets"))
        tables_selector = parse_selector(rule_row.get("Tables"))
        has_explicit_tables = bool(tables_selector)

        template_scope: Dict[str, Dict[str, Any]] = {}

        for template in templates:
            tables = self._parse_tables(rule_row.get("Tables"), template)
            available_sheets = self.repository.all_sheets(template)
            table_sheets: Dict[str, List[str]] = {}

            for table in tables:
                if sheets_override is not None:
                    selected = [sheet for sheet in sheets_override if sheet in available_sheets]
                    table_sheets[table] = selected
                    continue

                if table == template:
                    table_sheets[table] = available_sheets[:1] if available_sheets else []
                else:
                    table_sheets[table] = [self.repository.get_table_sheet(template, table)]

            template_scope[template] = {
                "tables": tables,
                "table_sheets": table_sheets,
                "rows": rows,
                "columns": columns,
                "alias_groups": combine_alias_groups(row_alias_groups, column_alias_groups),
            }

        return {"templates": templates, "template_scope": template_scope}


class FormulaParser:
    REF_PATTERN = re.compile(r"\{([^{}]+)\}(\(\*\))?")

    def parse(self, expression: str) -> ParsedExpression:
        if not expression or not isinstance(expression, str):
            raise NonDslRuleError("Empty expression")

        text = expression.strip()
        if text.lower().startswith("for each"):
            raise NonDslRuleError("Natural language rules are not supported")

        refs: Dict[str, RefSpec] = {}
        ref_counter = 0

        def _replace_ref(match: re.Match[str]) -> str:
            nonlocal ref_counter
            key = f"__ref_{ref_counter}"
            refs[key] = self._parse_ref_token(match.group(1), bool(match.group(2)))
            ref_counter += 1
            return key

        normalized = self.REF_PATTERN.sub(_replace_ref, text)
        normalized = self._normalize_operators(normalized)
        normalized = self._normalize_largest_sum(normalized)
        normalized = self._normalize_aggregations(normalized)

        try:
            parsed = ast.parse(normalized, mode="eval").body
        except SyntaxError as exc:
            raise RuleEngineError(f"Invalid formula syntax: {expression}") from exc

        return ParsedExpression(text=normalized, refs=refs, ast_root=parsed)

    def _normalize_operators(self, text: str) -> str:
        normalized = " ".join(text.replace("\n", " ").split())
        normalized = re.sub(r"(\S+)\s*!=\s*empty\b", r"is_not_empty(\1)", normalized, flags=re.IGNORECASE)
        normalized = re.sub(r"(\S+)\s*=\s*empty\b", r"is_empty_value(\1)", normalized, flags=re.IGNORECASE)
        normalized = re.sub(r"\bempty\b", "None", normalized, flags=re.IGNORECASE)
        normalized = re.sub(r"(?<![<>=!])=(?!=)", "==", normalized)
        return normalized

    def _normalize_aggregations(self, text: str) -> str:
        out = text
        for func in AGG_FUNCS:
            out = self._convert_bracket_call(out, func)
        return out

    def _normalize_largest_sum(self, text: str) -> str:
        pattern = re.compile(r"\bsum\s*\(", flags=re.IGNORECASE)
        result: List[str] = []
        cursor = 0

        while True:
            match = pattern.search(text, cursor)
            if not match:
                result.append(text[cursor:])
                break

            start = match.start()
            open_idx = match.end() - 1
            depth = 0
            close_idx = -1
            for idx in range(open_idx, len(text)):
                if text[idx] == "(":
                    depth += 1
                elif text[idx] == ")":
                    depth -= 1
                    if depth == 0:
                        close_idx = idx
                        break

            if close_idx < 0:
                raise RuleEngineError("Unclosed parenthesis for sum(...) expression")

            inner = text[open_idx + 1 : close_idx].strip()
            largest_match = re.fullmatch(
                r"(?is)(\d+)\s+largest\s+values\s+among\s*\((.*)\)",
                inner,
            )

            result.append(text[cursor:start])
            if largest_match is None:
                result.append(text[start : close_idx + 1])
            else:
                top_n = largest_match.group(1)
                values_expr = largest_match.group(2).strip()
                result.append(f"largest_sum({top_n}, {values_expr})")

            cursor = close_idx + 1

        return "".join(result)

    def _convert_bracket_call(self, text: str, func_name: str) -> str:
        pattern = re.compile(rf"\b{func_name}\s*\[")
        result = []
        cursor = 0

        while True:
            match = pattern.search(text, cursor)
            if not match:
                result.append(text[cursor:])
                break

            start = match.start()
            open_idx = match.end() - 1
            depth = 0
            close_idx = -1
            for idx in range(open_idx, len(text)):
                if text[idx] == "[":
                    depth += 1
                elif text[idx] == "]":
                    depth -= 1
                    if depth == 0:
                        close_idx = idx
                        break

            if close_idx < 0:
                raise RuleEngineError(f"Unclosed bracket aggregation for {func_name}")

            result.append(text[cursor:start])
            inner = text[open_idx + 1 : close_idx]
            result.append(f"{func_name}({inner})")
            cursor = close_idx + 1

        return "".join(result)

    def _parse_ref_token(self, raw_ref: str, wildcard: bool) -> RefSpec:
        template: Optional[str] = None
        table: Optional[str] = None
        row: Optional[str] = None
        row_alias: Optional[str] = None
        column: Optional[str] = None
        column_alias: Optional[str] = None
        sheet: Optional[str] = None
        include_sheet_pattern: Optional[str] = None
        exclude_sheets: List[str] = []

        parts = [chunk.strip() for chunk in raw_ref.split(",") if chunk.strip()]

        for part in parts:
            cleaned = part.strip()
            lowered = cleaned.lower()

            qualifier_match = re.fullmatch(r"((?:r|c)[a-z0-9_]+)\s+in\s+.+", lowered)
            if qualifier_match is not None:
                lowered = qualifier_match.group(1)
                cleaned = lowered

            if lowered.startswith("(") and lowered.endswith(")"):
                lowered = lowered[1:-1].strip()
                cleaned = cleaned[1:-1].strip()

            if "excluding" in lowered:
                left, right = [x.strip() for x in re.split(r"\bexcluding\b", lowered, maxsplit=1)]
                pattern = left if left else "s.*"
                pattern = re.escape(pattern).replace("NNN", r"\\d+").replace(r"\*", ".*")
                include_sheet_pattern = f"^{pattern}$"
                for token in re.split(r"[;\s]+", right):
                    if token:
                        exclude_sheets.append(token.lower())
                continue

            if re.fullmatch(r"[A-Z]\d{2}\.\d{2}\.[A-Z]+", cleaned.upper()):
                table = cleaned.upper()
                template = normalize_template_id(".".join(table.split(".")[:2]))
                continue

            if re.fullmatch(r"[A-Z]\d{2}\.\d{2}", cleaned.upper()):
                template = normalize_template_id(cleaned)
                continue

            if re.fullmatch(r"r\d{2,4}", lowered):
                norm = normalize_axis_code(lowered[1:])
                row = norm.zfill(4) if norm else lowered[1:]
                continue

            if re.fullmatch(r"r[a-z][a-z0-9_]*", lowered):
                row_alias = lowered
                continue

            if re.fullmatch(r"c\d{2,4}", lowered):
                norm = normalize_axis_code(lowered[1:])
                column = norm.zfill(4) if norm else lowered[1:]
                continue

            if re.fullmatch(r"c[a-z][a-z0-9_]*", lowered):
                column_alias = lowered
                continue

            if re.fullmatch(r"qx\d+", lowered):
                sheet = lowered
                continue

            if lowered.startswith("s"):
                escaped = re.escape(lowered).replace('NNN', r'\\d+').replace(r'\\*', '.*')
                include_sheet_pattern = f"^{escaped}$"

        return RefSpec(
            template=template,
            table=table,
            row=row,
            row_alias=row_alias,
            column=column,
            column_alias=column_alias,
            sheet=sheet,
            include_sheet_pattern=include_sheet_pattern,
            exclude_sheets=tuple(exclude_sheets),
            wildcard=wildcard,
            label=raw_ref.strip(),
        )


class ValueResolver:
    def __init__(self, repository: Any, scope: Dict[str, Any]):
        self.repository = repository
        self.scope = scope

    def _resolve_axis(
        self,
        base_value: Optional[str],
        override: Optional[str],
        alias_name: Optional[str],
        alias_map: Optional[Mapping[str, str]],
    ) -> Optional[str]:
        if alias_name is not None and alias_map is not None:
            alias_value = alias_map.get(alias_name.lower())
            if alias_value is not None:
                return alias_value
        if override is None:
            return base_value
        norm = normalize_axis_code(override)
        return norm.zfill(4) if norm is not None else override

    def _read_value(self, coord: Coordinate) -> Any:
        if coord.row is None or coord.column is None:
            return None

        context = self.repository.context(coord.template, coord.sheet)
        row_idx = context.row_map.get(coord.row)
        col_idx = context.col_map.get(coord.column)
        if row_idx is None or col_idx is None:
            return None
        return context.dataframe.loc[row_idx, col_idx]

    def _selected_sheets(self, template: str, table: str, base_sheet: str, ref: RefSpec) -> List[str]:
        if ref.sheet is not None:
            return [ref.sheet]

        available = self.repository.all_sheets(template)
        if ref.include_sheet_pattern is None:
            return [base_sheet]

        selected = [sheet for sheet in available if re.match(ref.include_sheet_pattern, sheet.lower())]
        if ref.exclude_sheets:
            excluded = set(ref.exclude_sheets)
            selected = [sheet for sheet in selected if sheet.lower() not in excluded]

        return selected or [base_sheet]

    def resolve_ref(
        self,
        base: Coordinate,
        ref: RefSpec,
        alias_map: Optional[Mapping[str, str]] = None,
    ) -> Any:
        template = ref.template or base.template
        table = ref.table or base.table

        if ref.table and ref.sheet is None and ref.include_sheet_pattern is None:
            default_sheet = self.repository.get_table_sheet(template, table)
        else:
            default_sheet = base.sheet

        sheets = self._selected_sheets(template, table, default_sheet, ref)

        row = self._resolve_axis(base.row, ref.row, ref.row_alias, alias_map)
        column = self._resolve_axis(base.column, ref.column, ref.column_alias, alias_map)

        expand_rows = ref.wildcard and row is None
        expand_cols = ref.wildcard and column is None

        values: List[Any] = []
        for sheet in sheets:
            context = self.repository.context(template, sheet)
            rows = list(context.row_map.keys()) if expand_rows else [row]
            cols = list(context.col_map.keys()) if expand_cols else [column]

            for row_code, col_code in product(rows, cols):
                values.append(
                    self._read_value(
                        Coordinate(
                            template=template,
                            table=table,
                            row=row_code,
                            column=col_code,
                            sheet=sheet,
                        )
                    )
                )

        if len(values) > 1:
            return values
        return values[0] if values else None


class AstEvaluator:
    def __init__(self, arithmetic_approach: str, tolerance: float = INTERVAL_TOLERANCE):
        self.exact_mode = str(arithmetic_approach).strip().lower() != "interval"
        self.tolerance = tolerance

    def evaluate(self, node: ast.AST, env: Dict[str, Any]) -> Any:
        if isinstance(node, ast.Constant):
            return node.value

        if isinstance(node, ast.Tuple):
            return tuple(self.evaluate(elt, env) for elt in node.elts)

        if isinstance(node, ast.List):
            return [self.evaluate(elt, env) for elt in node.elts]

        if isinstance(node, ast.Name):
            if node.id not in env:
                raise RuleEngineError(f"Unknown symbol: {node.id}")
            return env[node.id]

        if isinstance(node, ast.UnaryOp):
            value = self.evaluate(node.operand, env)
            if isinstance(node.op, ast.USub):
                number_value = to_number(value)
                return -number_value if number_value is not None else None
            if isinstance(node.op, ast.UAdd):
                return to_number(value)
            if isinstance(node.op, ast.Not):
                return not bool(value)
            raise RuleEngineError("Unsupported unary operation")

        if isinstance(node, ast.BinOp):
            left = self.evaluate(node.left, env)
            right = self.evaluate(node.right, env)
            return self._eval_binop(node.op, left, right)

        if isinstance(node, ast.BoolOp):
            values = [bool(self.evaluate(v, env)) for v in node.values]
            if isinstance(node.op, ast.And):
                return all(values)
            if isinstance(node.op, ast.Or):
                return any(values)
            raise RuleEngineError("Unsupported boolean operation")

        if isinstance(node, ast.Compare):
            left = self.evaluate(node.left, env)
            for op, comparator in zip(node.ops, node.comparators):
                right = self.evaluate(comparator, env)
                if not self._compare(op, left, right):
                    return False
                left = right
            return True

        if isinstance(node, ast.Call):
            if not isinstance(node.func, ast.Name):
                raise RuleEngineError("Unsupported function call")

            func_name = node.func.id
            args = [self.evaluate(arg, env) for arg in node.args]

            if func_name == "sum":
                return values_sum(args)
            if func_name == "largest_sum":
                if not args:
                    return None
                return values_largest_sum(args[0], args[1:])
            if func_name == "count":
                return values_count(args)
            if func_name == "max":
                return values_max(args)
            if func_name == "min":
                return values_min(args)
            if func_name == "length":
                return values_length(args[0] if args else None)
            if func_name == "is_empty_value":
                return is_empty(args[0] if args else None)
            if func_name == "is_not_empty":
                return not is_empty(args[0] if args else None)

            raise RuleEngineError(f"Unsupported function: {func_name}")

        raise RuleEngineError(f"Unsupported AST node: {type(node).__name__}")

    def _eval_binop(self, op: ast.operator, left: Any, right: Any) -> Any:
        left_num = to_number(left)
        right_num = to_number(right)

        if isinstance(op, ast.Add):
            if left_num is not None and right_num is not None:
                return left_num + right_num
            return f"{left}{right}"
        if isinstance(op, ast.Sub):
            if left_num is None and right_num is None:
                return 0.0
            return None if left_num is None or right_num is None else left_num - right_num
        if isinstance(op, ast.Mult):
            return None if left_num is None or right_num is None else left_num * right_num
        if isinstance(op, ast.Div):
            if left_num is None or right_num is None or right_num == 0:
                return None
            return left_num / right_num

        raise RuleEngineError("Unsupported binary operation")

    def _compare(self, operator: ast.cmpop, left: Any, right: Any) -> bool:
        left_num = to_number(left)
        right_num = to_number(right)

        if left_num is not None and right_num is not None:
            if isinstance(operator, ast.Eq):
                if self.exact_mode:
                    return left_num == right_num
                return abs(left_num - right_num) <= self.tolerance
            if isinstance(operator, ast.NotEq):
                if self.exact_mode:
                    return left_num != right_num
                return abs(left_num - right_num) > self.tolerance
            if isinstance(operator, ast.Gt):
                return left_num > right_num - (0 if self.exact_mode else self.tolerance)
            if isinstance(operator, ast.GtE):
                return left_num >= right_num - (0 if self.exact_mode else self.tolerance)
            if isinstance(operator, ast.Lt):
                return left_num < right_num + (0 if self.exact_mode else self.tolerance)
            if isinstance(operator, ast.LtE):
                return left_num <= right_num + (0 if self.exact_mode else self.tolerance)

        if isinstance(operator, ast.Eq):
            return left == right
        if isinstance(operator, ast.NotEq):
            return left != right
        if isinstance(operator, ast.Gt):
            try:
                return left > right
            except TypeError:
                return False
        if isinstance(operator, ast.GtE):
            try:
                return left >= right
            except TypeError:
                return False
        if isinstance(operator, ast.Lt):
            try:
                return left < right
            except TypeError:
                return False
        if isinstance(operator, ast.LtE):
            try:
                return left <= right
            except TypeError:
                return False
        if isinstance(operator, ast.In):
            try:
                return left in right
            except TypeError:
                return False
        if isinstance(operator, ast.NotIn):
            try:
                return left not in right
            except TypeError:
                return True

        raise RuleEngineError("Unsupported comparison operator")


class RuleEvaluator:
    def __init__(self, repository: Any, tolerance: float = INTERVAL_TOLERANCE):
        self.repository = repository
        self.tolerance = tolerance
        self.dimension_resolver = DimensionResolver(repository)
        self.formula_parser = FormulaParser()

    def _candidate_axis(
        self,
        context_map: Dict[str, int],
        selected: Optional[List[str]],
        anchor: Optional[str] = None,
        alias_map: Optional[Mapping[str, str]] = None,
        axis_prefix: Optional[str] = None,
    ) -> List[Optional[str]]:
        if selected is None:
            if anchor is not None and anchor in context_map:
                axis: List[Optional[str]] = [anchor]
            elif alias_map is not None and axis_prefix is not None:
                alias_values = [
                    value
                    for key, value in alias_map.items()
                    if key.startswith(axis_prefix.lower()) and value in context_map
                ]
                axis = list(dict.fromkeys(alias_values))
            else:
                axis = list(context_map.keys())
        else:
            axis = [code for code in selected if code in context_map]
        return axis or [None]

    def _resolve_ref_anchor(
        self,
        explicit_value: Optional[str],
        alias_name: Optional[str],
        alias_map: Optional[Mapping[str, str]],
    ) -> Optional[str]:
        if alias_name is not None and alias_map is not None:
            alias_value = alias_map.get(alias_name.lower())
            if alias_value is not None:
                return alias_value
        return explicit_value

    def evaluate_rule(self, rule_row: pd.Series) -> RuleResult:
        rule_id = str(rule_row.get("Id", "UNKNOWN"))
        formula = rule_row.get("Formula")
        precondition = rule_row.get("Precondition")
        arithmetic = str(rule_row.get("Arithmetic approach", "exact"))

        if formula is None or (isinstance(formula, float) and pd.isna(formula)):
            return RuleResult(rule_id=rule_id, status="SKIPPED", details=[], reason="Missing formula")

        try:
            scope = self.dimension_resolver.resolve_scope(rule_row)
            value_resolver = ValueResolver(self.repository, scope)
            evaluator = AstEvaluator(arithmetic_approach=arithmetic, tolerance=self.tolerance)
            formula_expr = self.formula_parser.parse(str(formula))
            pre_expr = (
                self.formula_parser.parse(str(precondition))
                if precondition is not None and not (isinstance(precondition, float) and pd.isna(precondition))
                else None
            )
        except Exception as exc:
            return RuleResult(rule_id=rule_id, status="SKIPPED", details=[], reason=str(exc))

        details: List[RuleDetail] = []
        any_fail = False

        first_formula_ref: Optional[RefSpec] = None
        if formula_expr.refs:
            first_formula_ref = next(iter(formula_expr.refs.values()))

        for template in scope["templates"]:
            meta = scope["template_scope"][template]
            for table in meta["tables"]:
                for sheet in meta["table_sheets"][table]:
                    context = self.repository.context(template, sheet)
                    alias_groups = meta.get("alias_groups") or [{}]

                    for alias_map in alias_groups:
                        anchor_row = self._resolve_ref_anchor(
                            first_formula_ref.row if first_formula_ref is not None else None,
                            first_formula_ref.row_alias if first_formula_ref is not None else None,
                            alias_map,
                        )
                        anchor_col = self._resolve_ref_anchor(
                            first_formula_ref.column if first_formula_ref is not None else None,
                            first_formula_ref.column_alias if first_formula_ref is not None else None,
                            alias_map,
                        )

                        row_candidates = self._candidate_axis(
                            context.row_map,
                            meta["rows"],
                            anchor=anchor_row,
                            alias_map=alias_map,
                            axis_prefix="r",
                        )
                        col_candidates = self._candidate_axis(
                            context.col_map,
                            meta["columns"],
                            anchor=anchor_col,
                            alias_map=alias_map,
                            axis_prefix="c",
                        )

                        for row_code, col_code in product(row_candidates, col_candidates):
                            coordinate = Coordinate(
                                template=template,
                                table=table,
                                row=row_code,
                                column=col_code,
                                sheet=sheet,
                            )

                            env: Dict[str, Any] = {"None": None, "True": True, "False": False}
                            formula_values: Dict[str, Any] = {}

                            try:
                                for ref_name, ref_spec in formula_expr.refs.items():
                                    resolved = value_resolver.resolve_ref(coordinate, ref_spec, alias_map=alias_map)
                                    env[ref_name] = resolved
                                    formula_values[f"{{{ref_spec.label}}}"] = resolved
                            except Exception as exc:
                                details.append(
                                    RuleDetail(
                                        coordinates=(
                                            coordinate.template,
                                            coordinate.table,
                                            coordinate.row,
                                            coordinate.column,
                                            coordinate.sheet,
                                        ),
                                        expected=True,
                                        actual=None,
                                        passed=False,
                                        message=f"Reference resolution error: {exc}",
                                        formula_values=formula_values,
                                    )
                                )
                                any_fail = True
                                continue

                            if pre_expr is not None:
                                pre_env = dict(env)
                                precondition_values: Dict[str, Any] = {}
                                try:
                                    for ref_name, ref_spec in pre_expr.refs.items():
                                        resolved_pre = value_resolver.resolve_ref(coordinate, ref_spec, alias_map=alias_map)
                                        pre_env[ref_name] = resolved_pre
                                        precondition_values[f"{{{ref_spec.label}}}"] = resolved_pre
                                except Exception as exc:
                                    details.append(
                                        RuleDetail(
                                            coordinates=(
                                                coordinate.template,
                                                coordinate.table,
                                                coordinate.row,
                                                coordinate.column,
                                                coordinate.sheet,
                                            ),
                                            expected=True,
                                            actual=None,
                                            passed=False,
                                            message=f"Precondition resolution error: {exc}",
                                            formula_values=formula_values,
                                            precondition_values=precondition_values,
                                        )
                                    )
                                    any_fail = True
                                    continue
                                try:
                                    if not bool(evaluator.evaluate(pre_expr.ast_root, pre_env)):
                                        continue
                                except Exception as exc:
                                    details.append(
                                        RuleDetail(
                                            coordinates=(
                                                coordinate.template,
                                                coordinate.table,
                                                coordinate.row,
                                                coordinate.column,
                                                coordinate.sheet,
                                            ),
                                            expected=True,
                                            actual=None,
                                            passed=False,
                                            message=f"Precondition evaluation error: {exc}",
                                            formula_values=formula_values,
                                            precondition_values=precondition_values,
                                        )
                                    )
                                    any_fail = True
                                    continue
                            else:
                                precondition_values = {}

                            try:
                                actual = evaluator.evaluate(formula_expr.ast_root, env)
                                passed = bool(actual)
                                message = ""
                            except Exception as exc:
                                actual = None
                                passed = False
                                message = f"Evaluation error: {exc}"

                            details.append(
                                RuleDetail(
                                    coordinates=(
                                        coordinate.template,
                                        coordinate.table,
                                        coordinate.row,
                                        coordinate.column,
                                        coordinate.sheet,
                                    ),
                                    expected=True,
                                    actual=actual,
                                    passed=passed,
                                    message=message,
                                    formula_values=formula_values,
                                    precondition_values=precondition_values,
                                )
                            )
                            if not passed:
                                any_fail = True

        if not details:
            return RuleResult(rule_id=rule_id, status="SKIPPED", details=[], reason="No coordinates evaluated")

        return RuleResult(rule_id=rule_id, status="FAIL" if any_fail else "PASS", details=details)


def _build_repository_from_mapping(data_mapping: Optional[Mapping[str, Any]], corep_dir: str | Path) -> Any:
    if data_mapping is not None:
        return InMemoryCorepDataRepository(data_mapping)
    return CorepDataRepository(corep_dir=corep_dir)


def evaluate_rule(
    rule_row: pd.Series,
    data_mapping: Optional[Mapping[str, Any]] = None,
    corep_dir: str | Path = DEFAULT_COREP_DIR,
) -> Dict[str, Any]:
    repository = _build_repository_from_mapping(data_mapping, corep_dir)
    return RuleEvaluator(repository).evaluate_rule(rule_row).to_dict()


def load_rules(
    config_path: str | Path = DEFAULT_BASED_TEMPLATE_PATH,
    sheet_name: str = DEFAULT_BASED_TEMPLATE_SHEET,
) -> pd.DataFrame:
    return pd.read_excel(config_path, sheet_name=sheet_name, header=1)


def evaluate_rules(
    config_path: str | Path = DEFAULT_BASED_TEMPLATE_PATH,
    sheet_name: str = DEFAULT_BASED_TEMPLATE_SHEET,
    corep_dir: str | Path = DEFAULT_COREP_DIR,
    max_rules: Optional[int] = None,
) -> List[Dict[str, Any]]:
    rules_df = load_rules(config_path=config_path, sheet_name=sheet_name)
    repository = CorepDataRepository(corep_dir=corep_dir)
    evaluator = RuleEvaluator(repository)

    results: List[Dict[str, Any]] = []
    for idx, (_, row) in enumerate(rules_df.iterrows(), start=1):
        results.append(evaluator.evaluate_rule(row).to_dict())
        if max_rules is not None and idx >= max_rules:
            break

    return results


if __name__ == "__main__":
    output = evaluate_rules(max_rules=10)
    summary = pd.Series([item["status"] for item in output]).value_counts().to_dict()
    print("Summary:", summary)
    print("Sample:", output[0] if output else {})
