"""
create_stub_workbooks.py

Generates minimal COREP workbook stubs for all templates referenced in the
config that do not yet have a corresponding file in COREP_files.

Each stub workbook contains one sheet per table suffix needed (a, b, c, …).
Each sheet embeds:
  - A header cell with the template code and table marker (so table discovery
    via resolve_sheet_for_table_generic works)
  - Row codes in column A
  - Column codes in row 2
  - Deterministic integer values at every row × column intersection
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
import pandas as pd

try:
    from v2.excel_io import read_excel_quiet
    from v2.get_RC_value import (
        DEFAULT_BASED_TEMPLATE_PATH,
        DEFAULT_BASED_TEMPLATE_SHEET,
        DEFAULT_COREP_DIR,
        letter_to_sequence,
        normalize_axis_code,
        parse_selector,
        sequence_to_marker,
    )
    from v2.rule_engine import normalize_template_id, template_to_file_name, RuleEngineError
except ModuleNotFoundError:
    from v2.excel_io import read_excel_quiet  # type: ignore
    from v2.get_RC_value import (  # type: ignore
        DEFAULT_BASED_TEMPLATE_PATH,
        DEFAULT_BASED_TEMPLATE_SHEET,
        DEFAULT_COREP_DIR,
        letter_to_sequence,
        normalize_axis_code,
        parse_selector,
        sequence_to_marker,
    )
    from v2.rule_engine import normalize_template_id, template_to_file_name, RuleEngineError  # type: ignore


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_codes(value: object) -> Set[str]:
    """Return normalised 4-digit axis codes from a cell value."""
    tokens = parse_selector(value) or []
    out: Set[str] = set()
    for token in tokens:
        norm = normalize_axis_code(token)
        if norm is not None:
            out.add(norm.zfill(4))
    return out


def _dsl_axis_codes(text: object) -> Tuple[Set[str], Set[str]]:
    """Extract row/column codes mentioned inside {…} DSL tokens."""
    rows: Set[str] = set()
    cols: Set[str] = set()
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return rows, cols
    for ref in re.findall(r"\{([^{}]+)\}", str(text)):
        for part in [chunk.strip().lower() for chunk in ref.split(",") if chunk.strip()]:
            if re.fullmatch(r"r\d{2,4}", part):
                norm = normalize_axis_code(part[1:])
                if norm:
                    rows.add(norm.zfill(4))
            elif re.fullmatch(r"c\d{2,4}", part):
                norm = normalize_axis_code(part[1:])
                if norm:
                    cols.add(norm.zfill(4))
    return rows, cols


def _table_suffixes_for_template(template: str, rules_df: pd.DataFrame) -> Set[str]:
    """Return table suffix letters (e.g. {'a', 'b'}) referenced for template."""
    suffixes: Set[str] = set()
    for _, row in rules_df.iterrows():
        templates_val = row.get("Templates used")
        templates_cell = [t.strip().upper() for t in str(templates_val).split(";") if t.strip()] if templates_val and not (isinstance(templates_val, float) and pd.isna(templates_val)) else []
        if template not in templates_cell:
            continue

        # From Tables column
        tables_val = row.get("Tables")
        if tables_val and not (isinstance(tables_val, float) and pd.isna(tables_val)):
            for tbl in str(tables_val).split(";"):
                tbl = tbl.strip().upper()
                parts = tbl.split(".")
                if len(parts) >= 3 and parts[-1].isalpha() and tbl.startswith(template):
                    suffixes.add(parts[-1].lower())

        # From formula / precondition DSL references to explicit table names
        for col in ("Formula", "Precondition"):
            val = row.get(col)
            if val and not (isinstance(val, float) and pd.isna(val)):
                for ref in re.findall(r"\{([^{}]+)\}", str(val)):
                    for part in [p.strip().upper() for p in ref.split(",")]:
                        m = re.fullmatch(r"[A-Z]\d{2}\.\d{2}\.([A-Z]+)", part)
                        if m:
                            suffixes.add(m.group(1).lower())

    return suffixes


def _axes_for_template(template: str, rules_df: pd.DataFrame) -> Tuple[Set[str], Set[str]]:
    """Return (row_codes, col_codes) for a template from all its rules."""
    rows: Set[str] = set()
    cols: Set[str] = set()
    for _, rule_row in rules_df.iterrows():
        templates_val = rule_row.get("Templates used")
        if not templates_val or (isinstance(templates_val, float) and pd.isna(templates_val)):
            continue
        templates = [t.strip().upper() for t in str(templates_val).split(";") if t.strip()]
        if template not in templates:
            continue
        rows.update(_parse_codes(rule_row.get("Rows")))
        cols.update(_parse_codes(rule_row.get("Columns")))
        for col_name in ("Formula", "Precondition"):
            r, c = _dsl_axis_codes(rule_row.get(col_name))
            rows.update(r)
            cols.update(c)
    return rows, cols


def _seed_value(template: str, sheet: str, row: str, col: str) -> int:
    seed = f"{template}|{sheet}|{row}|{col}"
    return (sum(ord(ch) for ch in seed) % 97) + 1


def _write_stub_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    template: str,
    marker_label: str,
    row_codes: List[str],
    col_codes: List[str],
    sheet_name: str,
) -> None:
    """
    Layout:
      A1 = "{template} - {marker}"  (enables table + template discovery)
      Row 2 = column codes starting at B2
      Col A from row 3 = row codes
      Intersections = seeded values
    """
    ws["A1"] = f"{template} - {marker_label}"
    for col_idx, col_code in enumerate(col_codes, start=2):
        ws.cell(row=2, column=col_idx).value = int(col_code)
    for row_idx, row_code in enumerate(row_codes, start=3):
        ws.cell(row=row_idx, column=1).value = int(row_code)
        for col_idx, col_code in enumerate(col_codes, start=2):
            ws.cell(row=row_idx, column=col_idx).value = _seed_value(template, sheet_name, row_code, col_code)


def create_stubs(
    config_path: Path = DEFAULT_BASED_TEMPLATE_PATH,
    sheet_name: str = DEFAULT_BASED_TEMPLATE_SHEET,
    corep_dir: Path = DEFAULT_COREP_DIR,
) -> Dict[str, int]:
    rules_df = read_excel_quiet(config_path, sheet_name=sheet_name, header=1)

    # Collect all referenced templates
    all_templates: Set[str] = set()
    for val in rules_df["Templates used"].dropna():
        for t in str(val).split(";"):
            t = t.strip()
            if not t:
                continue
            try:
                all_templates.add(normalize_template_id(t))
            except RuleEngineError:
                pass  # skip invalid codes silently

    corep_dir = Path(corep_dir)
    created = 0
    skipped = 0

    for template in sorted(all_templates):
        target = corep_dir / template_to_file_name(template)
        if target.exists():
            skipped += 1
            continue

        row_codes_set, col_codes_set = _axes_for_template(template, rules_df)
        # Always have at least some default codes so sheets are non-empty
        row_codes = sorted(row_codes_set or {"0010", "0020", "0030", "0040", "0050"})
        col_codes = sorted(col_codes_set or {"0010", "0020", "0030", "0040", "0050"})

        suffixes = _table_suffixes_for_template(template, rules_df) or {"a"}

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        for suffix in sorted(suffixes):
            marker_num = sequence_to_marker(letter_to_sequence(suffix))
            ws_name = f"{template}.{suffix}"[:31]  # Excel sheet names ≤ 31 chars
            ws = wb.create_sheet(title=ws_name)
            _write_stub_sheet(ws, template, marker_num, row_codes, col_codes, ws_name)

        # Also add a catch-all generic sheet for non-table rules
        ws_generic = wb.create_sheet(title="Sheet1")
        _write_stub_sheet(ws_generic, template, "0001", row_codes, col_codes, "Sheet1")

        wb.save(target)
        created += 1

    return {"created": created, "already_present": skipped, "total_templates": len(all_templates)}


def main() -> None:
    # ── PARAMETERS – edit here, then run ──────────────────────────────
    CONFIG    = DEFAULT_BASED_TEMPLATE_PATH
    SHEET     = DEFAULT_BASED_TEMPLATE_SHEET
    COREP_DIR = DEFAULT_COREP_DIR
    # ──────────────────────────────────────────────────────────────────

    result = create_stubs(config_path=CONFIG, sheet_name=SHEET, corep_dir=COREP_DIR)
    print(result)


if __name__ == "__main__":
    main()
