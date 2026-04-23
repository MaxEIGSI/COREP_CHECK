from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from string import ascii_lowercase
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

try:
    from v2.excel_io import load_workbook_quiet, read_excel_quiet
except ModuleNotFoundError:
    from v2.excel_io import load_workbook_quiet, read_excel_quiet  # type: ignore


MODULE_DIR = Path(__file__).resolve().parent
DATA_DIR = MODULE_DIR / "data"
DEFAULT_COREP_DIR = DATA_DIR / "COREP_files"
DEFAULT_BASED_TEMPLATE_PATH = DATA_DIR / "EGDQ_publication_2026.xlsx"
DEFAULT_BASED_TEMPLATE_SHEET = "v4.2"
DEFAULT_MAPPING_TABLE_PATH = DATA_DIR / "mapping_table.xlsx"
DEFAULT_QX_MAPPING_PATH   = DATA_DIR / "Mapping onglets COREP.xlsx"

ALL_SENTINEL = "__ALL__"


@dataclass(frozen=True)
class TableResolution:
    table_name: str
    sheet_name: str


class CorepExtractionError(Exception):
    pass


def resolve_corep_dir(corep_dir: Optional[str | Path] = None) -> Path:
    return Path(corep_dir) if corep_dir is not None else DEFAULT_COREP_DIR


def resolve_based_template_path(
    based_template_path: Optional[str | Path] = None,
) -> Path:
    return (
        Path(based_template_path)
        if based_template_path is not None
        else DEFAULT_BASED_TEMPLATE_PATH
    )


def resolve_based_template_sheet(sheet_name: Optional[str] = None) -> str:
    return sheet_name if sheet_name else DEFAULT_BASED_TEMPLATE_SHEET


def resolve_mapping_table_path(mapping_table_path: Optional[str | Path] = None) -> Path:
    return Path(mapping_table_path) if mapping_table_path is not None else DEFAULT_MAPPING_TABLE_PATH


def _normalize_table_key(value: Any) -> str:
    return str(value).strip().upper().replace(" ", "")


def _table_template_key(table_key: str) -> Optional[str]:
    match = re.match(r"([A-Z]\d{2}\.\d{2})", table_key)
    return match.group(1) if match else None


def _find_mapped_sheet_value(mapping: Dict[str, str], table_key: str) -> Optional[str]:
    direct = mapping.get(table_key)
    if direct is not None:
        return direct

    template_key = _table_template_key(table_key)
    if template_key is None:
        return None

    candidate_outputs = {
        output
        for key, output in mapping.items()
        if _table_template_key(key) == template_key
    }
    if len(candidate_outputs) == 1:
        return next(iter(candidate_outputs))

    return mapping.get(template_key)


def load_table_sheet_mapping(
    mapping_table_path: Optional[str | Path] = None,
) -> Dict[str, str]:
    path = resolve_mapping_table_path(mapping_table_path)
    if not path.exists():
        return {}

    df = read_excel_quiet(path)
    if df.empty:
        return {}

    normalized_cols = {col: str(col).strip().lower().replace(" ", "_") for col in df.columns}

    input_col = None
    output_col = None
    for col, norm in normalized_cols.items():
        if input_col is None and "input" in norm and "table" in norm:
            input_col = col
        if output_col is None and "output" in norm and ("table" in norm or "sheet" in norm):
            output_col = col

    if input_col is None or output_col is None:
        if len(df.columns) >= 2:
            input_col = df.columns[0]
            output_col = df.columns[1]
        else:
            return {}

    mapping: Dict[str, str] = {}
    for _, row in df.iterrows():
        table_in = row.get(input_col)
        sheet_out = row.get(output_col)
        if table_in is None or sheet_out is None:
            continue
        if pd.isna(table_in) or pd.isna(sheet_out):
            continue
        table_key = _normalize_table_key(table_in)
        sheet_name = str(sheet_out).strip()
        if table_key and sheet_name:
            mapping[table_key] = sheet_name

    return mapping


def find_mapped_sheet_for_table(
    table_name: str,
    mapping_table_path: Optional[str | Path] = None,
) -> Optional[str]:
    mapping = load_table_sheet_mapping(mapping_table_path)
    table_key = _normalize_table_key(table_name)
    return _find_mapped_sheet_value(mapping, table_key)


def _normalize_sheet_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value).upper())


def _match_sheet_name(wb: openpyxl.Workbook, mapped_name: str) -> Optional[str]:
    if mapped_name in wb.sheetnames:
        return mapped_name

    target = _normalize_sheet_key(mapped_name)
    if not target:
        return None

    for sheet in wb.sheetnames:
        if _normalize_sheet_key(sheet) == target:
            return sheet

    for sheet in wb.sheetnames:
        candidate = _normalize_sheet_key(sheet)
        if candidate.startswith(target) or target in candidate:
            return sheet

    return None


def parse_selector(value: Any) -> Optional[List[str]]:
    if value is None:
        return None

    if isinstance(value, float) and pd.isna(value):
        return None

    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        if stripped.lower() == "all":
            return [ALL_SENTINEL]
        return [item.strip() for item in stripped.split(";") if item.strip()]

    if isinstance(value, (list, tuple, set)):
        normalized = []
        for item in value:
            if item is None:
                continue
            item_str = str(item).strip()
            if not item_str:
                continue
            if item_str.lower() == "all":
                return [ALL_SENTINEL]
            normalized.append(item_str)
        return normalized or None

    return [str(value).strip()]


def normalize_template_code(template_code: str) -> str:
    code = template_code.strip().upper().replace(" ", "")
    if not re.fullmatch(r"C\d{2}\.\d{2}", code):
        raise CorepExtractionError(f"Invalid template code format: {template_code}")
    return code


def template_to_file_name(template_code: str) -> str:
    code = normalize_template_code(template_code)
    digits = code.replace("C", "").replace(".", "")
    return f"G_EU_C_C{digits}.xlsx"


def template_to_file_path(template_code: str, corep_dir: Path | str) -> Path:
    corep_path = resolve_corep_dir(corep_dir)
    file_path = corep_path / template_to_file_name(template_code)
    if not file_path.exists():
        raise CorepExtractionError(
            f"COREP file not found for template {template_code}: {file_path}"
        )
    return file_path


def split_table_name(table_name: str) -> Tuple[str, Optional[str]]:
    cleaned = table_name.strip().upper().replace(" ", "")
    parts = cleaned.split(".")
    if len(parts) < 2:
        return cleaned, None

    if len(parts) >= 3 and parts[-1].isalpha():
        template = ".".join(parts[:-1])
        return template, parts[-1].lower()

    return cleaned, None


def letter_to_sequence(letter: str) -> int:
    letter = letter.strip().lower()
    if not re.fullmatch(r"[a-z]+", letter):
        raise CorepExtractionError(f"Invalid table suffix letter: {letter}")

    value = 0
    for ch in letter:
        value = value * 26 + (ascii_lowercase.index(ch) + 1)
    return value


def sequence_to_marker(sequence: int) -> str:
    return f"{sequence:04d}"


def normalize_axis_code(value: Any) -> Optional[str]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    if re.fullmatch(r"\d+(\.0+)?", text):
        numeric_text = text.split(".")[0]
        return str(int(numeric_text))

    if re.fullmatch(r"\d{2,4}", text):
        return str(int(text))

    return None


def worksheet_strings(ws: Worksheet) -> Iterable[str]:
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if value is None:
                continue
            text = str(value).strip()
            if text:
                yield text


def contains_template_hint(text: str, template_code: str) -> bool:
    token1 = template_code.upper()
    token2 = template_code.upper().replace(".", "")
    token3 = token1.replace("C", "C ")
    up = text.upper()
    return token1 in up or token2 in up or token3 in up


def resolve_sheet_for_table(
    wb: openpyxl.Workbook,
    template_code: str,
    table_name: str,
    mapping_table_path: Optional[str | Path] = None,
) -> str:
    mapped_sheet = find_mapped_sheet_for_table(table_name, mapping_table_path)
    if mapped_sheet is not None:
        matched = _match_sheet_name(wb, mapped_sheet)
        if matched is not None:
            return matched

    template_from_table, suffix_letter = split_table_name(table_name)
    normalized_template = normalize_template_code(template_code)

    if normalize_template_code(template_from_table) != normalized_template:
        raise CorepExtractionError(
            f"Table {table_name} does not match template {template_code}"
        )

    if suffix_letter is None:
        raise CorepExtractionError(
            f"Table {table_name} has no suffix letter (expected format Cxx.xx.a)"
        )

    marker = sequence_to_marker(letter_to_sequence(suffix_letter))
    marker_pattern = re.compile(rf"-\s*{marker}\b")

    candidates: List[str] = []

    for ws in wb.worksheets:
        has_marker = False
        has_template = False

        for text in worksheet_strings(ws):
            if marker_pattern.search(text):
                has_marker = True
            if contains_template_hint(text, normalized_template):
                has_template = True
            if has_marker and has_template:
                candidates.append(ws.title)
                break

    if not candidates:
        raise CorepExtractionError(
            f"No worksheet found for table {table_name} (marker {marker})"
        )

    return candidates[0]


def worksheet_to_expanded_matrix(ws: Worksheet) -> List[List[Any]]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    matrix: List[List[Any]] = [[None for _ in range(max_col)] for _ in range(max_row)]

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.row is None or cell.column is None:
                continue
            matrix[int(cell.row) - 1][int(cell.column) - 1] = cell.value

    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left = ws.cell(row=min_row, column=min_col).value
        for r_idx in range(min_row - 1, max_row):
            for c_idx in range(min_col - 1, max_col):
                if matrix[r_idx][c_idx] is None:
                    matrix[r_idx][c_idx] = top_left

    return matrix


def worksheet_to_dataframe(ws: Worksheet) -> pd.DataFrame:
    matrix = worksheet_to_expanded_matrix(ws)
    return pd.DataFrame(matrix)


def build_row_code_map(df: pd.DataFrame, search_cols: int = 10) -> Dict[str, int]:
    row_map: Dict[str, int] = {}
    max_cols = min(search_cols, len(df.columns))

    for idx in df.index:
        row_values = [df.iat[idx, col] for col in range(max_cols)]
        for val in row_values:
            normalized_code = normalize_axis_code(val)
            if normalized_code is not None:
                row_map[normalized_code] = idx
                break

    return row_map


def build_column_code_map(df: pd.DataFrame, header_rows: int = 20) -> Dict[str, Any]:
    col_map: Dict[str, Any] = {}
    max_rows = min(header_rows, len(df.index))

    for col_position, col in enumerate(df.columns):
        for row_idx in range(max_rows):
            value = df.iat[row_idx, col_position]
            normalized_code = normalize_axis_code(value)
            if normalized_code is not None:
                col_map[normalized_code] = col
                break

    return col_map


def select_rows(df: pd.DataFrame, row_selector: Optional[List[str]]) -> pd.DataFrame:
    if not row_selector or ALL_SENTINEL in row_selector:
        return df

    row_map = build_row_code_map(df)
    indexes = []
    for token in row_selector:
        normalized_token = normalize_axis_code(token)
        if normalized_token is not None and normalized_token in row_map:
            indexes.append(row_map[normalized_token])
    if not indexes:
        return df.iloc[0:0, :]
    return df.loc[indexes, :]


def select_columns(df: pd.DataFrame, column_selector: Optional[List[str]]) -> pd.DataFrame:
    if not column_selector or ALL_SENTINEL in column_selector:
        return df

    col_map = build_column_code_map(df)
    columns = []
    for token in column_selector:
        normalized_token = normalize_axis_code(token)
        if normalized_token is not None and normalized_token in col_map:
            columns.append(col_map[normalized_token])
    if not columns:
        return df.iloc[:, 0:0]
    return df.loc[:, columns]


def filter_dataframe(
    df: pd.DataFrame,
    row_selector: Optional[List[str]],
    column_selector: Optional[List[str]],
) -> pd.DataFrame:
    filtered_df = df
    matched_row_codes: List[str] = []
    matched_column_codes: List[str] = []

    if row_selector and ALL_SENTINEL not in row_selector:
        row_map = build_row_code_map(df)
        indexes = []
        for token in row_selector:
            normalized_token = normalize_axis_code(token)
            if normalized_token is not None and normalized_token in row_map:
                indexes.append(row_map[normalized_token])
                matched_row_codes.append(normalized_token.zfill(4))
        filtered_df = filtered_df.loc[indexes, :] if indexes else filtered_df.iloc[0:0, :]

    if column_selector and ALL_SENTINEL not in column_selector:
        col_map = build_column_code_map(df)
        columns = []
        for token in column_selector:
            normalized_token = normalize_axis_code(token)
            if normalized_token is not None and normalized_token in col_map:
                columns.append(col_map[normalized_token])
                matched_column_codes.append(normalized_token.zfill(4))
        filtered_df = filtered_df.loc[:, columns] if columns else filtered_df.iloc[:, 0:0]

    if matched_row_codes and len(filtered_df.index) == len(matched_row_codes):
        filtered_df = filtered_df.copy()
        filtered_df.index = matched_row_codes

    if matched_column_codes and len(filtered_df.columns) == len(matched_column_codes):
        filtered_df = filtered_df.copy()
        filtered_df.columns = matched_column_codes

    return filtered_df


def parse_tables_for_template(
    template_code: str,
    tables: Optional[List[str]],
) -> Optional[List[str]]:
    if not tables:
        return None

    normalized = normalize_template_code(template_code)
    scoped: List[str] = []

    for table in tables:
        table = table.strip()
        if not table:
            continue

        table_template, _ = split_table_name(table)
        if normalize_template_code(table_template) == normalized:
            scoped.append(table)

    return scoped or None


def select_sheet_names(
    wb: openpyxl.Workbook,
    sheets: Optional[List[str]],
) -> List[str]:
    if not sheets or ALL_SENTINEL in sheets:
        return [ws.title for ws in wb.worksheets]

    available = {ws.title for ws in wb.worksheets}
    selected = [sheet for sheet in sheets if sheet in available]
    if not selected:
        raise CorepExtractionError(
            f"None of the requested sheets exist. Requested={sheets}, available={sorted(available)}"
        )
    return selected


def extract_table_dataframes(
    template_code: str,
    file_path: Path,
    tables: Optional[List[str]],
    rows: Optional[List[str]],
    columns: Optional[List[str]],
    sheets: Optional[List[str]],
    mapping_table_path: Optional[str | Path] = None,
) -> Dict[str, Dict[str, Any]]:
    wb = load_workbook_quiet(file_path, data_only=True)

    table_map: Dict[str, Dict[str, Any]] = {}

    scoped_tables = parse_tables_for_template(template_code, tables)

    if scoped_tables:
        for table_name in scoped_tables:
            sheet_name = resolve_sheet_for_table(
                wb,
                template_code,
                table_name,
                mapping_table_path=mapping_table_path,
            )
            ws = wb[sheet_name]
            df = filter_dataframe(worksheet_to_dataframe(ws), rows, columns)
            table_map[table_name] = {
                "sheet_name": sheet_name,
                "dataframe": df,
            }
        return table_map

    selected_sheets = select_sheet_names(wb, sheets)
    for sheet_name in selected_sheets:
        ws = wb[sheet_name]
        df = filter_dataframe(worksheet_to_dataframe(ws), rows, columns)
        table_map[sheet_name] = {
            "sheet_name": sheet_name,
            "dataframe": df,
        }

    return table_map


def get_value(
    templates_used: Any,
    tables: Any = None,
    rows: Any = None,
    columns: Any = None,
    sheets: Any = None,
    corep_dir: Optional[str | Path] = None,
    mapping_table_path: Optional[str | Path] = None,
) -> Dict[str, Dict[str, Any]]:
    template_selector = parse_selector(templates_used)
    if not template_selector:
        raise CorepExtractionError("templates_used is required")
    if ALL_SENTINEL in template_selector:
        raise CorepExtractionError("templates_used cannot be 'All'")

    table_selector = parse_selector(tables)
    row_selector = parse_selector(rows)
    column_selector = parse_selector(columns)
    sheet_selector = parse_selector(sheets)

    result: Dict[str, Dict[str, Any]] = {}
    resolved_corep_dir = resolve_corep_dir(corep_dir)

    for template_code in template_selector:
        normalized_template = normalize_template_code(template_code)
        file_path = template_to_file_path(normalized_template, resolved_corep_dir)

        table_map = extract_table_dataframes(
            template_code=normalized_template,
            file_path=file_path,
            tables=table_selector,
            rows=row_selector,
            columns=column_selector,
            sheets=sheet_selector,
            mapping_table_path=mapping_table_path,
        )

        result[normalized_template] = {
            "file_path": str(file_path),
            "tables": table_map,
        }

    return result


def load_based_template(
    based_template_path: Optional[str | Path] = None,
    sheet_name: Optional[str] = None,
) -> pd.DataFrame:
    resolved_path = resolve_based_template_path(based_template_path)
    resolved_sheet = resolve_based_template_sheet(sheet_name)
    return read_excel_quiet(resolved_path, sheet_name=resolved_sheet)


def run_from_based_template_row(
    row: pd.Series,
    corep_dir: Optional[str | Path] = None,
) -> Dict[str, Dict[str, Any]]:
    return get_value(
        templates_used=row.get("Templates used"),
        tables=row.get("Tables"),
        rows=row.get("Rows"),
        columns=row.get("Columns"),
        sheets=row.get("Sheets"),
        corep_dir=corep_dir,
    )


def run_from_based_template(
    based_template_path: Optional[str | Path] = None,
    sheet_name: Optional[str] = None,
    corep_dir: Optional[str | Path] = None,
) -> List[Dict[str, Dict[str, Any]]]:
    based_df = load_based_template(based_template_path, sheet_name)
    outputs: List[Dict[str, Dict[str, Any]]] = []

    for _, row in based_df.iterrows():
        try:
            outputs.append(run_from_based_template_row(row, corep_dir=corep_dir))
        except CorepExtractionError:
            continue

    return outputs


# if __name__ == "__main__":
#     examples = [
#         {
#             "label": "Requested example",
#             "params": {
#                 "templates_used": "C07.00",
#                 "tables": "C07.00.a",
#                 "rows": "0010;0020",
#                 "columns": "0010;0020",
#             },
#         },
#         {
#             "label": "Non-empty example",
#             "params": {
#                 "templates_used": "C07.00",
#                 "tables": "C07.00.b",
#                 "rows": "240",
#                 "columns": "240",
#             },
#         },
#     ]

#     for example in examples:
#         print(f"\n=== {example['label']} ===")
#         sample = get_value(**example["params"])

#         for template, template_payload in sample.items():
#             print(f"Template: {template}")
#             print(f"File: {template_payload['file_path']}")
#             for table_name, table_payload in template_payload["tables"].items():
#                 print(f"  Table: {table_name} -> Sheet: {table_payload['sheet_name']}")
#                 print(f"  Data shape: {table_payload['dataframe'].shape}")
#                 print(table_payload["dataframe"].to_string())
