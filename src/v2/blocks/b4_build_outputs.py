"""
Block 4 – Build outputs
=======================
Writes the evaluation results to an Excel workbook (Summary + Details sheets)
and a JSON backup.

Input context keys:  results (List[Dict]), rules_df (pd.DataFrame),
                     output_dir (Path)
Output context keys: evaluate_summary  {"total_rules", "status_counts",
                                         "xlsx_path", "json_path"}
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd


# ---------------------------------------------------------------------------
# Config columns forwarded to Summary sheet for easy inspection
# ---------------------------------------------------------------------------
_CONFIG_COLS = [
    "Id", "Status", "Severity", "Templates used", "Tables", "Rows", "Columns",
    "Sheets", "Precondition", "Formula", "Arithmetic approach", "Description",
]


def _clean(val: Any) -> Any:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return val


def _to_json_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    try:
        return json.dumps(value, ensure_ascii=False, default=str)
    except Exception:
        return str(value)


def _serialize_coordinate(coordinates: Any) -> Dict[str, Any]:
    coords = coordinates if isinstance(coordinates, (list, tuple)) else ()
    return {
        "template": coords[0] if len(coords) > 0 else None,
        "table": coords[1] if len(coords) > 1 else None,
        "row": coords[2] if len(coords) > 2 else None,
        "column": coords[3] if len(coords) > 3 else None,
        "sheet": coords[4] if len(coords) > 4 else None,
    }


def _collect_all_values(details: List[Dict[str, Any]], value_key: str) -> List[Dict[str, Any]]:
    collected: List[Dict[str, Any]] = []
    for detail in details:
        collected.append(
            {
                "coordinates": _serialize_coordinate(detail.get("coordinates")),
                "passed": detail.get("passed"),
                "actual": detail.get("actual"),
                "message": detail.get("message") or "",
                "values": detail.get(value_key, {}) or {},
            }
        )
    return collected


def _collect_all_traces(details: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    traces: List[Dict[str, Any]] = []
    for detail in details:
        traces.append(
            {
                "coordinates": _serialize_coordinate(detail.get("coordinates")),
                "passed": detail.get("passed"),
                "trace": detail.get("evaluation_trace") or "",
            }
        )
    return traces


def _summarize_rule(result: Dict[str, Any], config_row: pd.Series) -> Dict[str, Any]:
    details = result.get("details", [])
    fail_count = sum(1 for d in details if not d.get("passed", False))
    row: Dict[str, Any] = {col: _clean(config_row.get(col)) for col in _CONFIG_COLS}
    row["Engine status"] = result.get("status")
    row["Skip / error reason"] = result.get("reason") or ""
    row["Evaluated points"] = len(details)
    row["Failed points"] = fail_count
    row["All evaluation traces"] = _to_json_text(_collect_all_traces(details))
    row["All formula values"] = _to_json_text(_collect_all_values(details, "formula_values"))
    row["All precondition values"] = _to_json_text(_collect_all_values(details, "precondition_values"))
    return row


def _flatten_details(result: Dict[str, Any]) -> List[Dict[str, Any]]:
    rule_id = result.get("rule_id")
    status = result.get("status")
    rows = []
    for d in result.get("details", []):
        coords = d.get("coordinates", ())
        rows.append({
            "Rule Id": rule_id,
            "Rule status": status,
            "Template": coords[0] if len(coords) > 0 else None,
            "Table": coords[1] if len(coords) > 1 else None,
            "Row": coords[2] if len(coords) > 2 else None,
            "Column": coords[3] if len(coords) > 3 else None,
            "Sheet": coords[4] if len(coords) > 4 else None,
            "Expected": d.get("expected"),
            "Actual": d.get("actual"),
            "Passed": d.get("passed"),
            "Evaluation trace": d.get("evaluation_trace") or "",
            "Formula values": _to_json_text(d.get("formula_values", {})),
            "Precondition values": _to_json_text(d.get("precondition_values", {})),
            "Message": d.get("message") or "",
        })
    return rows


def _style_summary(ws: Any, n_rules: int) -> None:
    from openpyxl.styles import PatternFill, Font
    RED    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    GREEN  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    BOLD   = Font(bold=True)

    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=False)))
    for cell in header:
        cell.font = BOLD

    status_col = next((c.column for c in header if c.value == "Engine status"), None)
    if status_col is None:
        return

    fill_map = {"PASS": GREEN, "FAIL": RED, "SKIPPED": YELLOW}
    for row_idx in range(2, n_rules + 2):
        cell = ws.cell(row=row_idx, column=status_col)
        fill = fill_map.get(cell.value)
        if fill:
            cell.fill = fill


def block_build_outputs(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """Write evaluation results to rule_results.xlsx and rule_results.json."""
    results   = ctx["results"]
    rules_df  = ctx.get("rules_df", pd.DataFrame())
    output_dir = Path(ctx.get("output_dir", "src/v2/data/output"))
    output_dir.mkdir(parents=True, exist_ok=True)

    config_by_id: Dict[str, pd.Series] = {
        str(row.get("Id", "")): row
        for _, row in rules_df.iterrows()
        if row.get("Id") is not None
    }

    summary_rows: List[Dict[str, Any]] = []
    detail_rows: List[Dict[str, Any]] = []
    for result in results:
        rule_id = result.get("rule_id", "")
        config_row = config_by_id.get(rule_id, pd.Series(dtype=object))
        summary_rows.append(_summarize_rule(result, config_row))
        detail_rows.extend(_flatten_details(result))

    summary_df = pd.DataFrame(summary_rows)
    details_df = pd.DataFrame(detail_rows)

    xlsx_path = output_dir / "rule_results.xlsx"
    with pd.ExcelWriter(str(xlsx_path), engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        details_df.to_excel(writer, sheet_name="Details", index=False)

        ws_summary = writer.sheets["Summary"]
        for col_cells in ws_summary.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            ws_summary.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)
        _style_summary(ws_summary, len(summary_rows))

        ws_details = writer.sheets["Details"]
        for col_cells in ws_details.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            ws_details.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

    json_path = output_dir / "rule_results.json"
    json_path.write_text(
        json.dumps(results, indent=2, default=str, ensure_ascii=False), encoding="utf-8"
    )

    counts = summary_df["Engine status"].value_counts().to_dict() if not summary_df.empty else {}
    ctx["evaluate_summary"] = {
        "total_rules": len(results),
        "status_counts": counts,
        "xlsx_path": str(xlsx_path),
        "json_path": str(json_path),
    }
    return ctx
