from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd

try:
    from v2.get_RC_value import DEFAULT_BASED_TEMPLATE_PATH, DEFAULT_BASED_TEMPLATE_SHEET, DEFAULT_COREP_DIR
    from v2.rule_engine import evaluate_rules, load_rules
except ModuleNotFoundError:
    from v2.get_RC_value import DEFAULT_BASED_TEMPLATE_PATH, DEFAULT_BASED_TEMPLATE_SHEET, DEFAULT_COREP_DIR  # type: ignore
    from v2.rule_engine import evaluate_rules, load_rules  # type: ignore

# Config columns to carry into the output for easy inspection
_CONFIG_COLS = [
    "Id", "Status", "Severity", "Templates used", "Tables", "Rows", "Columns",
    "Sheets", "Precondition", "Formula", "Arithmetic approach", "Description",
]


def _clean(val: Any) -> Any:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return val


def _summarize_rule(result: Dict[str, Any], config_row: pd.Series) -> Dict[str, Any]:
    details = result.get("details", [])
    fail_count = sum(1 for d in details if not d.get("passed", False))
    row: Dict[str, Any] = {col: _clean(config_row.get(col)) for col in _CONFIG_COLS}
    row["Engine status"] = result.get("status")
    row["Skip / error reason"] = result.get("reason") or ""
    row["Evaluated points"] = len(details)
    row["Failed points"] = fail_count
    return row


def _flatten_details(result: Dict[str, Any]) -> List[Dict[str, Any]]:
    rule_id = result.get("rule_id")
    status = result.get("status")
    rows = []
    for d in result.get("details", []):
        coords = d.get("coordinates", ())
        rows.append(
            {
                "Rule Id": rule_id,
                "Rule status": status,
                "Template": coords[0] if len(coords) > 0 else None,
                "Table": coords[1] if len(coords) > 1 else None,
                "Row": coords[2] if len(coords) > 2 else None,
                "Column": coords[3] if len(coords) > 3 else None,
                "Sheet": coords[4] if len(coords) > 4 else None,
                "Expected": d.get("expected"),
                "Actual": d.get("actual"),
                "Passed": d.get("passed"),
                "Message": d.get("message") or "",
            }
        )
    return rows


def _style_summary(ws: Any, n_rules: int) -> None:
    from openpyxl.styles import PatternFill, Font

    RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    BOLD = Font(bold=True)

    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=False)))
    for cell in header:
        cell.font = BOLD

    status_col = next((c.column for c in header if c.value == "Engine status"), None)
    if status_col is None:
        return

    for row_idx in range(2, n_rules + 2):
        cell = ws.cell(row=row_idx, column=status_col)
        if cell.value == "PASS":
            cell.fill = GREEN
        elif cell.value == "FAIL":
            cell.fill = RED
        elif cell.value == "SKIPPED":
            cell.fill = YELLOW


def run_pipeline(
    config_path: Path = DEFAULT_BASED_TEMPLATE_PATH,
    sheet_name: str = DEFAULT_BASED_TEMPLATE_SHEET,
    corep_dir: Path = DEFAULT_COREP_DIR,
    output_dir: Path = Path("outputs"),
    max_rules: int | None = None,
) -> Dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)

    config_df = load_rules(config_path=config_path, sheet_name=sheet_name)

    results = evaluate_rules(
        config_path=config_path,
        sheet_name=sheet_name,
        corep_dir=corep_dir,
        max_rules=max_rules,
    )

    config_by_id: Dict[str, pd.Series] = {
        str(row.get("Id", "")): row
        for _, row in config_df.iterrows()
        if row.get("Id") is not None
    }

    summary_rows: List[Dict[str, Any]] = []
    detail_rows: List[Dict[str, Any]] = []

    for result in results:
        rule_id = result.get("rule_id", "")
        config_row = config_by_id.get(rule_id, pd.Series(dtype=object))
        summary_rows.append(_summarize_rule(result, config_row))
        detail_rows.extend(_flatten_details(result))

    summary_df = pd.DataFrame(summary_rows)
    details_df = pd.DataFrame(detail_rows)

    xlsx_path = output_dir / "rule_results.xlsx"
    with pd.ExcelWriter(str(xlsx_path), engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        details_df.to_excel(writer, sheet_name="Details", index=False)

        ws_summary = writer.sheets["Summary"]
        for col_cells in ws_summary.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            ws_summary.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)
        _style_summary(ws_summary, len(summary_rows))

        ws_details = writer.sheets["Details"]
        for col_cells in ws_details.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            ws_details.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

    # Lightweight JSON backup
    json_path = output_dir / "rule_results.json"
    json_path.write_text(
        json.dumps(results, indent=2, default=str, ensure_ascii=False), encoding="utf-8"
    )

    counts = summary_df["Engine status"].value_counts().to_dict() if not summary_df.empty else {}
    return {
        "total_rules": len(results),
        "status_counts": counts,
        "xlsx_path": str(xlsx_path),
        "json_path": str(json_path),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Run COREP rule pipeline and export PASS/FAIL per row")
    parser.add_argument("--config", type=Path, default=DEFAULT_BASED_TEMPLATE_PATH)
    parser.add_argument("--sheet", type=str, default=DEFAULT_BASED_TEMPLATE_SHEET)
    parser.add_argument("--corep-dir", type=Path, default=DEFAULT_COREP_DIR)
    parser.add_argument("--output-dir", type=Path, default=Path("outputs"))
    parser.add_argument("--max-rules", type=int, default=None)
    args = parser.parse_args()

    result = run_pipeline(
        config_path=args.config,
        sheet_name=args.sheet,
        corep_dir=args.corep_dir,
        output_dir=args.output_dir,
        max_rules=args.max_rules,
    )
    print(result)


if __name__ == "__main__":
    main()
